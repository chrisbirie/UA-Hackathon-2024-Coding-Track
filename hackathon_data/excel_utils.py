from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def create_excel_workbook(filename, data):
    # Create a new workbook
    workbook = Workbook()

    # Add a worksheet
    worksheet = workbook.active
    worksheet.title = "Data"

    # Write headers
    headers = ["Username", "Full Name", "Exercise", "Languages", "Ended At", "Rank", "Score"]
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, row_data in enumerate(data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    # Autosize columns
    for column in worksheet.columns:
        max_length = 0
        column = list(column)  # Convert to list to enable re-iterations over it
        for cell in column:
            # Make sure to load the cell's value as a string
            cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)

        adjusted_width = (max_length + 2)  # Adding a little extra space
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(filename)

    return workbook