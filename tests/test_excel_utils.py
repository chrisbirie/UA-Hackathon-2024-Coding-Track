import unittest
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from hackathon_data.excel_utils import create_excel_workbook

class TestCreateExcelWorkbook(unittest.TestCase):
    def test_create_excel_workbook(self):
        # Define the test file name
        test_filename = 'test_output.xlsx'

        # Mock data
        data = [
            ["user1", "John Doe", "Exercise1", "Python", "2024-04-18 12:00:00", "8 kyu", 50],
            ["user2", "Jane Smith", "Exercise2", "Java", "2024-04-18 13:00:00", "7 kyu", 40]
        ]

        # Call the function
        workbook = create_excel_workbook(test_filename, data)

        # Assertions
        self.assertTrue(os.path.exists(test_filename))  # Check if the file was created

        # Read the Excel file using openpyxl
        workbook = load_workbook(filename=test_filename)
        sheet = workbook.active

        # Ensure headers are correct
        headers = ["Username", "Full Name", "Exercise", "Languages", "Ended At", "Rank", "Score"]
        for col_num, header in enumerate(headers, start=1):
            self.assertEqual(sheet.cell(row=1, column=col_num).value, header)

        # Ensure data is correct
        for row_num, row_data in enumerate(data, start=2):
            for col_num, expected_value in enumerate(row_data, start=1):
                self.assertEqual(sheet.cell(row=row_num, column=col_num).value, expected_value)

        # Remove the test file
        os.remove(test_filename)

class TestCreateExcelWorkbook(unittest.TestCase):
    def test_column_widths(self):
        data = [
            ['user12345', 'User One Very Long Name', 'Exercise A', 'Python', '2024-04-01', 'Beginner', 100],
            ['user2', 'User Two', 'Exercise B', 'Java, Scala, Kotlin', '2024-04-02', 'Intermediate', 200]
        ]
        test_filename = 'test_output.xlsx'
        
        # Call the function
        workbook = create_excel_workbook(test_filename, data)
        worksheet = workbook.active

        # Get the maximum content length for each column, adjusted by logic used in function
        expected_widths = []
        for column in worksheet.columns:
            max_length = 0
            for cell in column:
                max_length = max(max_length, len(str(cell.value)))
            expected_widths.append(max_length + 2)

        # Check if each column width matches the expected width
        for i, column in enumerate(worksheet.columns, start=1):
            column_letter = get_column_letter(i)
            self.assertEqual(worksheet.column_dimensions[column_letter].width, expected_widths[i-1])


if __name__ == "__main__":
    unittest.main()